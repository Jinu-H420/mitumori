#!/usr/bin/env python3
"""
曲げ加工見積り 計算書（Excel）を生成する。
Googleスプレッドシートで開いて使う想定。
実行: python3 build_calc_sheet.py
要: pip install openpyxl
"""
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    raise

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")
OUT_PATH = os.path.join(SCRIPT_DIR, "曲げ加工見積り_計算書.xlsx")


def load_data():
    with open(os.path.join(DATA_DIR, "prices.json"), encoding="utf-8") as f:
        prices = json.load(f)
    with open(os.path.join(DATA_DIR, "rates.json"), encoding="utf-8") as f:
        rates = json.load(f)
    hole_path = os.path.join(DATA_DIR, "hole_prices.json")
    hole_prices = None
    if os.path.exists(hole_path):
        with open(hole_path, encoding="utf-8") as f:
            hole_prices = json.load(f)
    return prices, rates, hole_prices


def main():
    prices, rates, hole_prices = load_data()
    wb = Workbook()
    thin = Side(style="thin")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold = Font(bold=True)

    # =============================================
    # シート「データ」: 全マスターデータ
    # =============================================
    ws_data = wb.active
    ws_data.title = "データ"

    # --- 基準価格テーブル (A1〜F12) ---
    weight_limits = prices["重量区分"]
    length_limits = prices["長さ区分"]
    price_table = prices["基準価格"]

    weight_labels = []
    for w in weight_limits:
        if w >= 99999:
            weight_labels.append(f">{weight_limits[-2]}kg")
        else:
            weight_labels.append(f"≤{w}kg")

    length_labels = []
    for ll in length_limits:
        if ll >= 99999:
            length_labels.append(f">{length_limits[-2]}mm")
        else:
            length_labels.append(f"≤{ll}mm")

    ws_data["A1"] = "基準価格(L曲げ)"
    ws_data["A1"].font = Font(bold=True, size=11)
    for j, label in enumerate(length_labels):
        c = ws_data.cell(row=2, column=j + 2, value=label)
        c.font = bold
        c.fill = header_fill
        c.border = border_all
    for i, row_data in enumerate(price_table):
        c = ws_data.cell(row=i + 3, column=1, value=weight_labels[i])
        c.font = bold
        c.fill = header_fill
        c.border = border_all
        for j, val in enumerate(row_data):
            c = ws_data.cell(row=i + 3, column=j + 2, value=val)
            c.border = border_all
            c.number_format = "#,##0"

    n_weights = len(weight_labels)  # 11
    n_lengths = len(length_labels)  # 5
    # 基準価格テーブル範囲: B3:F13

    # --- 重量区分・長さ区分（MATCH参照用）(H列) ---
    ws_data["H1"] = "重量区分(kg)"
    ws_data["H1"].font = bold
    for i, w in enumerate(weight_limits):
        ws_data.cell(row=i + 2, column=8, value=w)

    ws_data["I1"] = "長さ区分(mm)"
    ws_data["I1"].font = bold
    for i, ll in enumerate(length_limits):
        ws_data.cell(row=i + 2, column=9, value=ll)

    # --- 形状乗率 (A16〜) ---
    shape_start = n_weights + 5  # row 16
    ws_data.cell(row=shape_start, column=1, value="形状").font = bold
    ws_data.cell(row=shape_start, column=2, value="乗率").font = bold
    ws_data.cell(row=shape_start, column=1).fill = header_fill
    ws_data.cell(row=shape_start, column=2).fill = header_fill
    shape_items = list(rates["形状乗率"].items())
    for i, (shape, r) in enumerate(shape_items):
        ws_data.cell(row=shape_start + 1 + i, column=1, value=shape).border = border_all
        ws_data.cell(row=shape_start + 1 + i, column=2, value=r).border = border_all

    # --- 数量調整 (D16〜) ---
    ws_data.cell(row=shape_start, column=4, value="数量min").font = bold
    ws_data.cell(row=shape_start, column=5, value="数量max").font = bold
    ws_data.cell(row=shape_start, column=6, value="調整率").font = bold
    for col in [4, 5, 6]:
        ws_data.cell(row=shape_start, column=col).fill = header_fill
    for i, band in enumerate(rates["数量調整"]):
        ws_data.cell(row=shape_start + 1 + i, column=4, value=band["min"]).border = border_all
        ws_data.cell(row=shape_start + 1 + i, column=5, value=band["max"] if band["max"] is not None else 9999).border = border_all
        ws_data.cell(row=shape_start + 1 + i, column=6, value=band["rate"]).border = border_all

    # --- 比重 (H16〜) ---
    ws_data.cell(row=shape_start, column=8, value="材質").font = bold
    ws_data.cell(row=shape_start, column=9, value="比重").font = bold
    ws_data.cell(row=shape_start, column=8).fill = header_fill
    ws_data.cell(row=shape_start, column=9).fill = header_fill
    mat_items = [(k, v) for k, v in rates["比重"].items() if v is not None]
    # CP400はnullなので除外し、別途「縞板」セクションで管理
    for i, (mat, g) in enumerate(mat_items):
        ws_data.cell(row=shape_start + 1 + i, column=8, value=mat).border = border_all
        ws_data.cell(row=shape_start + 1 + i, column=9, value=g).border = border_all
    # CP400行（比重欄に「単位重量表」と注記）
    cp_row = shape_start + 1 + len(mat_items)
    ws_data.cell(row=cp_row, column=8, value="CP400").border = border_all
    ws_data.cell(row=cp_row, column=9, value="単位重量表").border = border_all

    # --- 縞板単位重量 (H21〜) ---
    stripe_start = cp_row + 2
    ws_data.cell(row=stripe_start, column=8, value="縞板 板厚(mm)").font = bold
    ws_data.cell(row=stripe_start, column=9, value="単位重量(kg/m²)").font = bold
    ws_data.cell(row=stripe_start, column=8).fill = header_fill
    ws_data.cell(row=stripe_start, column=9).fill = header_fill
    stripe_data = rates.get("縞板_単位重量", {})
    stripe_items = [(k, v) for k, v in stripe_data.items() if k != "_doc"]
    for i, (t, w) in enumerate(stripe_items):
        ws_data.cell(row=stripe_start + 1 + i, column=8, value=float(t)).border = border_all
        ws_data.cell(row=stripe_start + 1 + i, column=9, value=w).border = border_all

    # --- 穴あけ単価 (A28〜) ---
    if hole_prices:
        hole_start = stripe_start + len(stripe_items) + 3
        ws_data.cell(row=hole_start, column=1, value="穴あけ単価").font = Font(bold=True, size=11)
        hr = hole_start + 1
        headers_h = ["板厚min", "板厚max", "丸穴_標準", "丸穴_下限", "長穴_標準", "長穴_下限"]
        for j, h in enumerate(headers_h):
            c = ws_data.cell(row=hr, column=j + 1, value=h)
            c.font = bold
            c.fill = header_fill
            c.border = border_all
        for i, row in enumerate(hole_prices["穴あけ単価"]):
            r = hr + 1 + i
            ws_data.cell(row=r, column=1, value=row["板厚min_mm"]).border = border_all
            ws_data.cell(row=r, column=2, value=row["板厚max_mm"]).border = border_all
            for j, key in enumerate(["丸穴_標準", "丸穴_下限", "長穴_標準", "長穴_下限"]):
                v = row.get(key)
                c = ws_data.cell(row=r, column=j + 3, value=v if v is not None else "")
                c.border = border_all

    # 列幅調整
    ws_data.column_dimensions["A"].width = 12
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws_data.column_dimensions[col_letter].width = 10
    ws_data.column_dimensions["H"].width = 16
    ws_data.column_dimensions["I"].width = 16

    # =============================================
    # シート「見積り」: 入力 → 自動計算
    # =============================================
    ws = wb.create_sheet("見積り", 0)

    # タイトル
    ws["A1"] = "曲げ加工 見積り計算書"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    # 入力エリア
    ws["A3"] = "【入力】"
    ws["A3"].font = Font(bold=True, size=11, color="2F5496")

    input_labels = [
        (4, "形状"),
        (5, "板厚(mm)"),
        (6, "展開幅(mm)"),
        (7, "製品長さ(mm)"),
        (8, "材質"),
        (9, "数量"),
        (10, "丸穴数"),
        (11, "長穴数"),
    ]
    for row_num, label in input_labels:
        c = ws.cell(row=row_num, column=1, value=label)
        c.font = bold
        c.border = border_all
        bc = ws.cell(row=row_num, column=2)
        bc.fill = input_fill
        bc.border = border_all

    # 初期値
    ws["B4"] = "L曲げ"
    ws["B5"] = 3.2
    ws["B6"] = 200
    ws["B7"] = 1500
    ws["B8"] = "SS400"
    ws["B9"] = 1
    ws["B10"] = 0
    ws["B11"] = 0

    # ドロップダウン: 形状
    shape_list = ",".join(rates["形状乗率"].keys())
    dv_shape = DataValidation(type="list", formula1=f'"{shape_list}"', allow_blank=False)
    dv_shape.error = "一覧から選択してください"
    ws.add_data_validation(dv_shape)
    dv_shape.add("B4")

    # ドロップダウン: 材質
    mat_list = ",".join([k for k in rates["比重"].keys()])
    dv_mat = DataValidation(type="list", formula1=f'"{mat_list}"', allow_blank=False)
    ws.add_data_validation(dv_mat)
    dv_mat.add("B8")

    # 計算エリア
    ws["A13"] = "【計算結果】"
    ws["A13"].font = Font(bold=True, size=11, color="2F5496")

    # セル参照のための変数
    # データシートの基準価格: B3:F13 (11行×5列)
    price_range = f"データ!$B$3:$F${2 + n_weights}"
    # 重量区分: H2:H12
    wt_range = f"データ!$H$2:$H${1 + n_weights}"
    # 長さ区分: I2:I6
    lt_range = f"データ!$I$2:$I${1 + n_lengths}"
    # 形状乗率VLOOKUP: A17:B23
    shape_range = f"データ!$A${shape_start + 1}:$B${shape_start + len(shape_items)}"
    # 数量調整VLOOKUP: D17:F19
    qty_range = f"データ!$D${shape_start + 1}:$F${shape_start + len(rates['数量調整'])}"
    # 比重VLOOKUP: H17:I18 (SS400, SUS304のみ)
    grav_range = f"データ!$H${shape_start + 1}:$I${shape_start + len(mat_items)}"
    # 縞板単位重量VLOOKUP: H(stripe_start+1):I(stripe_start+len)
    stripe_range = f"データ!$H${stripe_start + 1}:$I${stripe_start + len(stripe_items)}"

    # 穴あけデータ範囲（動的に計算）
    if hole_prices:
        n_holes = len(hole_prices["穴あけ単価"])
        hole_data_start = hole_start + 2  # hr+1
        hole_data_end = hole_data_start + n_holes - 1
        # A=板厚min, B=板厚max, C=丸穴_標準, E=長穴_標準
        hole_min_range = f"データ!$A${hole_data_start}:$A${hole_data_end}"
        hole_max_range = f"データ!$B${hole_data_start}:$B${hole_data_end}"
        hole_maru_range = f"データ!$C${hole_data_start}:$C${hole_data_end}"
        hole_naga_range = f"データ!$E${hole_data_start}:$E${hole_data_end}"
        hole_formula = (
            f'=IFERROR(SUMPRODUCT(({hole_min_range}<=B5)*({hole_max_range}>=B5)*{hole_maru_range})*B10,0)'
            f'+IFERROR(SUMPRODUCT(({hole_min_range}<=B5)*({hole_max_range}>=B5)*{hole_naga_range})*B11,0)'
        )
    else:
        hole_formula = "=0"

    calc_rows = [
        (14, "比重/単位重量", f'=IF(B8="CP400",IFERROR(VLOOKUP(B5,{stripe_range},2,0),"板厚なし"),IFERROR(VLOOKUP(B8,{grav_range},2,0),""))'),
        (15, "重量(kg)", f'=IF(B8="CP400",(B6/1000)*(B7/1000)*B14,(B6/1000)*(B7/1000)*(B5/1000)*(B14*1000))'),
        (16, "重量区分", f"=IFERROR(MATCH(B15,{wt_range},1),1)"),
        (17, "長さ区分", f"=IFERROR(MATCH(B7,{lt_range},1),1)"),
        (18, "基準価格", f"=INDEX({price_range},B16,B17)"),
        (19, "形状乗率", f'=IFERROR(VLOOKUP(B4,{shape_range},2,0),"")'),
        (20, "数量調整", f'=IFERROR(VLOOKUP(B9,{qty_range},3,1),"")'),
        (21, "単価(税抜)", '=IF(MOD(B18*B19*B20,50)=0,B18*B19*B20,ROUND(B18*B19*B20/100,0)*100)'),
        (22, "穴あけ加算", hole_formula),
        (23, "曲げ工賃(税抜)", "=B21*B9+B22"),
        (24, "消費税10%", "=ROUND(B23*0.1,0)"),
        (25, "合計(税込)", "=B23+B24"),
    ]

    for row_num, label, formula in calc_rows:
        c = ws.cell(row=row_num, column=1, value=label)
        c.font = bold
        c.border = border_all
        fc = ws.cell(row=row_num, column=2)
        fc.value = formula
        fc.border = border_all
        fc.number_format = "#,##0"

    # 単価と合計は目立たせる
    for r in [21, 23, 25]:
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2).font = Font(bold=True, size=11)
    ws.cell(row=25, column=2).font = Font(bold=True, size=12, color="C00000")

    # 列幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10

    wb.save(OUT_PATH)
    print(f"作成しました: {OUT_PATH}")


if __name__ == "__main__":
    main()
